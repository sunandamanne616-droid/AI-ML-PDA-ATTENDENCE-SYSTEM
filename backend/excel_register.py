"""
Excel Register for AIML Attendance System.
Matches the format shown in the department screenshot exactly:
  Row 1     : "Department of Artificial Intelligence and Machine Learning" — navy title
  Rows 2-6  : Info table  (Subject / Code / Semester / Dept / Faculty / Room / Session / Year / Generated / Total Classes)
  Row 7     : Stats banner (Total Students | Dates Covered)
  Row 8     : Column headers (Roll No | Name | Section | DD-Mon dates | Total | Present | Absent | %)
  Rows 9+   : Student data — alternating white/light-blue rows
              P = bold green text on light green bg
              A = bold red   text on light pink  bg
              - = plain grey text, white bg
              % cell colour: green ≥75%, orange 60-74%, red <60%
  Last row  : Class average footer

update_subject_excel() signature is UNCHANGED — existing calls work as-is.
Optional keyword args let callers pass faculty/session/room info.
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

EXCEL_ROOT = Path(__file__).resolve().parent / "excels"

# ── Colours (matched from screenshot) ─────────────────────
_NAVY     = "1B3A6B"   # title / header / banner background
_INFO_LBL = "EBF5FB"   # info label cell background (very light blue)
_INFO_VAL = "FFFFFF"   # info value cell background (white)
_ALT_ROW  = "EBF5FB"   # alternating data row (same light blue)
_WHITE    = "FFFFFF"
_P_BG     = "D5F5E3"   # present cell background (light green)
_P_FG     = "1E8449"   # present text (dark green)
_A_BG     = "FADBD8"   # absent cell background (light pink/red)
_A_FG     = "C0392B"   # absent text (dark red)
_DASH_FG  = "95A5A6"   # dash/missing text (grey)
_PCT_GRN  = "1E8449"   # % ≥ 75
_PCT_ORG  = "D35400"   # % 60-74
_PCT_RED  = "C0392B"   # % < 60
_FOOTER   = "1B3A6B"   # footer background

def _safe(s: str) -> str:
    return "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in s).strip().replace(" ", "_")

def get_subject_excel_path(semester: int, subject: str) -> Path:
    d = EXCEL_ROOT / f"SEM_{semester}"
    d.mkdir(parents=True, exist_ok=True)
    return d / f"{_safe(subject)}.xlsx"

def _thin(color="BDC3C7") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _styled(ws, row, col, value="", *,
            bold=False, size=10, fg="1A1A1A", bg=None,
            halign="center", valign="center", wrap=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=size, bold=bold, color=fg)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    c.border = _thin()
    if num_fmt:
        c.number_format = num_fmt
    return c

# ═══════════════════════════════════════════════════════════
# PUBLIC API — signature unchanged
# ═══════════════════════════════════════════════════════════

def update_subject_excel(
    semester: int,
    subject: str,
    students: List[dict],
    day: date,
    statuses: Dict[str, str],       # roll_no -> "P" / "A"
    # optional enrichment — all default to empty so existing calls work
    faculty_name: str = "",
    subject_code: str = "",
    session_time: str = "",
    room: str = "",
    dept: str = "AIML",
) -> Path:
    """
    Rebuild the full Excel register after each attendance session.
    Existing call in attendance.py requires no changes.
    """
    path = get_subject_excel_path(semester, subject)

    # ── 1. Read back all history from existing file ────────
    existing: Dict[str, Dict[str, str]] = {}   # roll -> {date_iso: P/A}
    existing_dates: List[str] = []

    if path.exists():
        try:
            old = load_workbook(path, data_only=True)
            ws0 = old.active
            hdr_row = None
            for r in range(1, min(ws0.max_row + 1, 15)):
                if ws0.cell(r, 1).value == "Roll No":
                    hdr_row = r
                    break
            if hdr_row:
                header = [ws0.cell(hdr_row, c).value for c in range(1, ws0.max_column + 1)]
                for h in header:
                    if isinstance(h, str) and len(h) == 10 and h[4] == "-":
                        existing_dates.append(h)
                    elif isinstance(h, date):
                        existing_dates.append(h.isoformat())
                for r in range(hdr_row + 1, ws0.max_row + 1):
                    roll = str(ws0.cell(r, 1).value or "").strip()
                    if not roll or roll == "CLASS AVERAGE":
                        continue
                    existing[roll] = {}
                    for ci, h in enumerate(header):
                        dk = None
                        if isinstance(h, str) and len(h) == 10 and h[4] == "-":
                            dk = h
                        elif isinstance(h, date):
                            dk = h.isoformat()
                        if dk:
                            v = ws0.cell(r, ci + 1).value
                            if v in ("P", "A"):
                                existing[roll][dk] = v
            old.close()
        except Exception:
            pass

    # ── 2. Merge today into history ────────────────────────
    today_iso = day.isoformat()
    for roll, status in statuses.items():
        if roll not in existing:
            existing[roll] = {}
        existing[roll][today_iso] = status

    all_dates = sorted(set(existing_dates + [today_iso]))
    stud_map  = {str(s["roll_no"]): s for s in students}
    all_rolls = sorted(set(list(stud_map.keys()) + list(existing.keys())))

    # ── 3. Build workbook ──────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Register"

    NF = 3                    # Roll No, Name, Section
    ND = len(all_dates)
    NS = 4                    # Total, Present, Absent, %
    TC = NF + ND + NS
    LC = get_column_letter(TC)

    # ── Row 1 : Title ─────────────────────────────────────
    ws.merge_cells(f"A1:{LC}1")
    _styled(ws, 1, 1,
            "Department of Artificial Intelligence and Machine Learning",
            bold=True, size=14, fg="FFFFFF", bg=_NAVY)
    ws.row_dimensions[1].height = 28

    # ── Rows 2-6 : Info block ─────────────────────────────
    #   Columns split roughly in half: left pair | right pair
    mid = TC // 2
    info = [
        ("Subject",   subject,                         "Subject Code",  subject_code or subject[:8].upper()),
        ("Semester",  f"Semester {semester}",          "Department",    dept),
        ("Faculty",   faculty_name or "—",             "Room",          room or "AIML-LAB-1"),
        ("Session",   session_time or "—",             "Academic Year", datetime.now().strftime("%Y-%m")),
        ("Generated", datetime.now().strftime("%d %b %Y %I:%M %p"),
                                                        "Total Classes", str(ND)),
    ]
    for idx, (l1, v1, l2, v2) in enumerate(info, start=2):
        # left label
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=2)
        _styled(ws, idx, 1, l1, bold=True, bg=_INFO_LBL, halign="left")
        # left value
        ws.merge_cells(start_row=idx, start_column=3, end_row=idx, end_column=mid)
        _styled(ws, idx, 3, v1, bg=_INFO_VAL, halign="left")
        # right label
        ws.merge_cells(start_row=idx, start_column=mid+1, end_row=idx, end_column=mid+2)
        _styled(ws, idx, mid+1, l2, bold=True, bg=_INFO_LBL, halign="left")
        # right value
        ws.merge_cells(start_row=idx, start_column=mid+3, end_row=idx, end_column=TC)
        _styled(ws, idx, mid+3, v2, bg=_INFO_VAL, halign="left")
        ws.row_dimensions[idx].height = 18

    # ── Row 7 : Stats banner ──────────────────────────────
    BANNER = 7
    ws.merge_cells(f"A{BANNER}:{LC}{BANNER}")
    date_range = f"{all_dates[0]} to {all_dates[-1]}" if all_dates else "—"
    _styled(ws, BANNER, 1,
            f"Total Students: {len(all_rolls)}    |    Dates Covered: {date_range}",
            bold=True, size=10, fg="FFFFFF", bg=_NAVY)
    ws.row_dimensions[BANNER].height = 18

    # ── Row 8 : Column headers ────────────────────────────
    HDR = 8
    for ci, label in enumerate(["Roll No", "Name", "Section"], start=1):
        _styled(ws, HDR, ci, label, bold=True, fg="FFFFFF", bg=_NAVY, size=10)

    for di, d_iso in enumerate(all_dates):
        col = NF + 1 + di
        try:
            dt = date.fromisoformat(d_iso)
            label = f"{dt.strftime('%d-%b')}\n{dt.strftime('%a')}"
        except Exception:
            label = d_iso
        _styled(ws, HDR, col, label, bold=True, fg="FFFFFF", bg=_NAVY, size=9, wrap=True)

    for si, sh in enumerate(["Total", "Present", "Absent", "%"]):
        _styled(ws, HDR, NF + ND + 1 + si, sh, bold=True, fg="FFFFFF", bg=_NAVY, size=10)
    ws.row_dimensions[HDR].height = 32

    # ── Rows 9+ : Student data ────────────────────────────
    DS = HDR + 1   # data start row

    for ri, roll in enumerate(all_rolls):
        row    = DS + ri
        stud   = stud_map.get(roll, {})
        name   = stud.get("name", "—")
        section = stud.get("section", "A") or "A"
        att    = existing.get(roll, {})
        present = sum(1 for v in att.values() if v == "P")
        total   = sum(1 for v in att.values() if v in ("P", "A"))
        absent  = total - present
        pct     = round((present / total) * 100, 1) if total > 0 else 0.0

        row_bg = _ALT_ROW if ri % 2 == 0 else _WHITE

        _styled(ws, row, 1, roll,    bg=row_bg, halign="left")
        _styled(ws, row, 2, name,    bg=row_bg, halign="left")
        _styled(ws, row, 3, section, bg=row_bg)

        for di, d_iso in enumerate(all_dates):
            col    = NF + 1 + di
            status = att.get(d_iso, "-")
            c = ws.cell(row=row, column=col, value=status)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin()
            if status == "P":
                c.font = Font(name="Arial", bold=True, color=_P_FG, size=10)
                c.fill = PatternFill("solid", start_color=_P_BG)
            elif status == "A":
                c.font = Font(name="Arial", bold=True, color=_A_FG, size=10)
                c.fill = PatternFill("solid", start_color=_A_BG)
            else:
                c.font = Font(name="Arial", color=_DASH_FG, size=9)
                c.fill = PatternFill("solid", start_color=row_bg)

        _styled(ws, row, NF + ND + 1, total,  bg=row_bg)
        _styled(ws, row, NF + ND + 2, present, bg=row_bg)
        _styled(ws, row, NF + ND + 3, absent,  bg=row_bg)

        # % cell — coloured text only, no full-row highlighting
        pct_fg = _PCT_GRN if pct >= 75 else (_PCT_ORG if pct >= 60 else _PCT_RED)
        c = ws.cell(row=row, column=NF + ND + 4, value=pct / 100 if total > 0 else 0)
        c.number_format = "0.0%"
        c.font      = Font(name="Arial", bold=True, size=10, color=pct_fg)
        c.fill      = PatternFill("solid", start_color=row_bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _thin()
        ws.row_dimensions[row].height = 16

    # ── Footer : class average ────────────────────────────
    footer = DS + len(all_rolls)
    ws.merge_cells(start_row=footer, start_column=1, end_row=footer, end_column=NF + ND)
    _styled(ws, footer, 1, "CLASS AVERAGE", bold=True, fg="FFFFFF", bg=_FOOTER)

    if all_rolls:
        avgs = []
        for roll in all_rolls:
            att = existing.get(roll, {})
            tot = sum(1 for v in att.values() if v in ("P", "A"))
            pre = sum(1 for v in att.values() if v == "P")
            if tot > 0:
                avgs.append(pre / tot)
        if avgs:
            avg_val = sum(avgs) / len(avgs)
            pct_fg  = _PCT_GRN if avg_val >= 0.75 else (_PCT_ORG if avg_val >= 0.60 else _PCT_RED)
            ac = ws.cell(footer, NF + ND + 4, avg_val)
            ac.number_format = "0.0%"
            ac.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            ac.fill      = PatternFill("solid", start_color=_FOOTER)
            ac.alignment = Alignment(horizontal="center", vertical="center")
            ac.border    = _thin()
    ws.row_dimensions[footer].height = 18

    # ── Column widths ──────────────────────────────────────
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 9
    for i in range(ND):
        ws.column_dimensions[get_column_letter(NF + 1 + i)].width = 7
    for i in range(NS):
        ws.column_dimensions[get_column_letter(NF + ND + 1 + i)].width = 10

    # ── Freeze Roll No / Name / Section ───────────────────
    ws.freeze_panes = ws.cell(row=DS, column=NF + 1)

    # ── Print layout ──────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1

    wb.save(path)
    return path

